import lpi
import openpyxl

def find_the_smallest(column, datatype, analysis_type):
    species_sizes = []
    for row in range(2, sheet.max_row):
        species_name = sheet.cell(row=row, column=1).value
        body_size = sheet.cell(row=row, column=column).value
        if body_size == None:
            continue
        measurement_unit = sheet.cell(row=row, column=column+1).value
        body_size = lpi.tidy_size(body_size, measurement_unit, datatype)
        if body_size == False:
            continue
        if analysis_type == "slope":
            abundance = float(sheet.cell(row=row, column=2).value)
        elif analysis_type == "percent_change":
            LPI_sheet = openpyxl.load_workbook(
                "Living Planet Index-07-04-2016.xlsx").active
            for LPI_row in range(2, LPI_sheet.max_row):
                if LPI_sheet.cell(row=LPI_row, column=2).value == species_name:
                    abundance = lpi.calculate_percent_changes(LPI_sheet, LPI_row)
                    break
        else:
            raise Exception("Invalid analysis type")
        data = [body_size, abundance]
        species_sizes.append(data)
    sorted_species = sorted(species_sizes, key=lambda species: species[0])
    bottom_ten_percent = len(sorted_species) / 10
    ten_percent = []
    for num in range(0, len(sorted_species)):
        if num < bottom_ten_percent:
            ten_percent.append(sorted_species[num][1])
        else:
            break
    return ten_percent

wb = openpyxl.load_workbook("tidy_lpi_data.xlsx")
sheet = wb.active
smallest_mass = find_the_smallest(3, "weight", "percent_change")
smallest_VT = find_the_smallest(5, "length", "percent_change")
smallest_CMO = find_the_smallest(7, "length", "percent_change")
print len(smallest_mass)
print lpi.mean(smallest_mass)
print len(smallest_VT)
print lpi.mean(smallest_VT)
print len(smallest_CMO)
print lpi.mean(smallest_CMO)
