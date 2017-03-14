import csv

import openpyxl
import lpi

read_wb = openpyxl.load_workbook(
    "Living Planet Index-07-04-2016.xlsx")
LPI_sheet = read_wb.active
species_info = {}

def mean(numbers):
    return float(sum(numbers) / max(len(numbers), 1))

def find_species_size_info(file_name, species_name, size_type):
    with open(file_name, "rb") as f:
        reader = csv.reader(f)
        species_info.setdefault(species_name, {})
        species_info[species_name].setdefault(size_type, None)
        for eol_row in reader:
            test_species_name = eol_row[1].decode('utf-8')
            if test_species_name == species_name:
                body_size = eol_row[4]
                measurement_unit = eol_row[7]
                if size_type == "Mass":
                    measurement_type = "weight"
                elif size_type == "Body Length VT" or size_type == "Body Length CMO":
                    measurement_type = "length"
                body_size = lpi.tidy_size(body_size, measurement_unit, measurement_type)
                species_info[species_name][size_type] = body_size
                break

for row in range(2, 401):
    species_name = LPI_sheet.cell(row=row, column=2).value
    #find_species_size_info("eol_download_2379.csv", species_name, "Mass")
    #find_species_size_info("eol_download_2416.csv", species_name, "Body Length VT")
    find_species_size_info("eol_download_2419.csv", species_name, "Body Length CMO")
    #slope_abundance = lpi.calculate_abundance_slopes(LPI_sheet, row)
    #species_info[species_name]["Slope Abundance"] = slope_abundance
    percent_change_abundance = lpi.calculate_percent_changes(LPI_sheet, row)
    species_info[species_name]["Percent Change Abundance"] = percent_change_abundance
    print(species_name)

write_wb = openpyxl.load_workbook("tidy_lpi_data.xlsx")
sheet = write_wb.active
for row in range(2, 400):
    species_name = sheet.cell(row=row, column=1).value
    sheet.cell(row=row, column=3).value = species_info[species_name]["Percent Change Abundance"]
    sheet.cell(row=row, column=6).value = species_info[species_name]["Body Mass CMO"]
write_wb.save("tidy_lpi_data_cleaner.xlsx")

# write_wb = openpyxl.Workbook()
# sheet = write_wb.active
# columns = ["Slope Abundance", "Percent Change Abundance", "Mass", "Body Length VT",
#     "Body Length CMO"]
# sheet['A1'] = "Species Name"
# column_number = 2
# for column in columns:
#    sheet.cell(row=1, column=column_number).value = column
#    column_number +=1
# row_number = 2
# for species in species_info:
#     print species + "making workbook"
#     sheet.cell(row=row_number, column=1).value = species
#     for num in range(2, 6):
#         column_header = columns[num - 2]
#         sheet.cell(row=row_number, column=num).value = species_info[species][column_header]
#     row_number += 1
# write_wb.save("tidy_lpi_data.xlsx")
