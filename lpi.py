import csv

import openpyxl
import math
from scipy import stats
import matplotlib.pyplot as plt

def create_plot(body_sizes, abundances, analysis_type):
    body_sizes.sort()
    if analysis_type == "big":
        top_x = body_sizes[len(body_sizes) - 1]
    elif analysis_type == "small":
        top_x = body_sizes[len(body_sizes) / 2]
    else:
        raise Exception("invalid analysis type")
    plt.plot(body_sizes, abundances, "o")
    plt.show()

def mean(numbers):
    return float(sum(numbers) / max(len(numbers), 1))

def adjust_weight(body_size, measurement_unit):
    if measurement_unit == "kg":
        return body_size * 1000
    elif measurement_unit == "log10 grams":
        return 10**body_size
    elif measurement_unit == "g":
        return body_size
    else:
        return False

def adjust_length(body_size, measurement_unit):
      if measurement_unit == "cm":
          return body_size * 10
      elif measurement_unit == "m":
          return body_size * 1000
      elif measurement_unit == "inch":
          return body_size * 25.4
      elif measurement_unit == "mm":
          return body_size
      else:
          return False

def tidy_size(body_size, measurement_unit, measurement_type):
    body_size = body_size.replace(",","")
    body_size = float(body_size)
    if measurement_type == "weight":
        body_size = adjust_weight(body_size, measurement_unit)
    elif measurement_type == "length":
        body_size = adjust_length(body_size, measurement_unit)
    else:
        raise Exception("Invalid measurement type")
    return body_size

def years_and_abundances(LPI_sheet, row):
    column = 3
    years = []
    abundances = []
    while (LPI_sheet.cell(row=row, column=column).value != None):
        year = LPI_sheet.cell(row=row, column=column).value
        abundance = LPI_sheet.cell(row=row,column=column+1).value
        years.append(float(year))
        abundances.append(float(abundance))
        column = column + 2
    return [years, abundances]

def calculate_abundance_slopes(LPI_sheet, row):
    years, abundances = years_and_abundances(LPI_sheet, row)
    log_abundances = []
    for abundance in abundances:
        log_abundances.append(math.asinh(abundance))
    abnormal_slope, _, _, _, _ = stats.linregress(years, log_abundances)
    abundances_mean = mean(abundances)
    slope = abnormal_slope / abundances_mean * 100
    return float(slope)

def calculate_percent_changes(LPI_sheet, row):
    years, abundances = years_and_abundances(LPI_sheet, row)
    annual_percent_changes = []
    mean_abundances = mean(abundances)
    for num in range(1, len(abundances)-1):
        difference = abundances[num] - abundances[num-1]
        percent_change = difference / mean_abundances * 100
        years_between = years[num] - years[num-1]
        annual_percent_change = percent_change / years_between
        annual_percent_changes.append(annual_percent_change)
    return mean(annual_percent_changes)

def analyze_LPI(eol_file, measurement_type, analysis_type):
    wb = openpyxl.load_workbook(
        "Living Planet Index-07-04-2016.xlsx")
    LPI_sheet = wb.active
    body_sizes= []
    changes = []
    counter = 0
    for row in range(2, 401):
        species_name = LPI_sheet.cell(row=row, column=2).value
        print species_name
        with open(eol_file, "rb") as f:
            reader = csv.reader(f)
            for eol_row in reader:
                test_species_name = eol_row[1].decode('utf-8')
                if test_species_name == species_name:
                    counter += 1
                    body_size = tidy_size(eol_row[4], eol_row[7], measurement_type)
                    if body_size == False:
                        break
                    body_sizes.append(body_size)
                    if analysis_type == "slopes":
                        changes.append(calculate_abundance_slopes(LPI_sheet, row))
                        break
                    elif analysis_type == "percent change":
                        changes.append(calculate_percent_changes(LPI_sheet, row))
                        break
                    else:
                        raise Exception ("Invalid analysis type")
    return [counter, body_sizes, changes]
