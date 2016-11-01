import csv

import openpyxl
import math
import matplotlib.pyplot as plt
from scipy import stats

def mean(numbers):
    return float(sum(numbers) / max(len(numbers), 1))

wb = openpyxl.load_workbook(
    "Living Planet Index-07-04-2016.xlsx")
LPI_sheet = wb.active
body_sizes= []
slopes = []
counter = 0
for row in range(2, 401):
    species_name = LPI_sheet.cell(row=row, column=2).value
#replace "eol_download_2419.csv" with whatever eol file you're using
    with open("eol_download_2379.csv", "rb") as f:
        reader = csv.reader(f)
        for eol_row in reader:
            test_species_name = eol_row[1].decode('utf-8')
            if test_species_name == species_name:
                counter += 1
                body_size = eol_row[4]
                body_size = body_size.replace(",","")
                body_size = float(body_size)
                measurement_unit = eol_row[7]
#comment this code in and the other set of ifs out if you're working with body mass
                if measurement_unit == "kg":
                    body_size = body_size * 1000
                elif measurement_unit == "log10 grams":
                    body_size = 10**body_size
                elif measurement_unit != "g":
                    break
#                if measurement_unit == "cm":
#                    body_size = body_size * 10
#                elif measurement_unit == "m":
#                    body_size = body_size * 1000
#                elif measurement_unit == "inch":
#                    body_size = body_size * 25.4
#                elif measurement_unit != "mm":
#                    break
                body_sizes.append(body_size)
                column = 3
                years = []
                abundances = []
                while (LPI_sheet.cell(row=row, column=column).value != None):
                    year = LPI_sheet.cell(row=row, column=column).value
                    abundance = LPI_sheet.cell(row=row,column=column+1).value
                    years.append(float(year))
                    abundances.append(float(abundance))
                    column = column + 2
                log_abundances = []
                for abundance in abundances:
                    log_abundances.append(math.asinh(abundance))
                abnormal_slope, _, _, _, _ = stats.linregress(years, log_abundances)
                abundances_mean = mean(abundances)
                slope = abnormal_slope / abundances_mean * 100
                slopes.append(float(slope))
                break
print counter
print stats.linregress(body_sizes, slopes)
plt.plot(body_sizes, slopes, "o")
plt.xlim([0, 5000])
plt.ylim([200, -200])
plt.show()
