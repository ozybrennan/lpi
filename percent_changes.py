import csv

import openpyxl
from scipy import stats

def mean(numbers):
    return float(sum(numbers) / max(len(numbers), 1))

wb = openpyxl.load_workbook(
    "Living Planet Index-07-04-2016.xlsx")
LPI_sheet = wb.active
body_sizes= []
average_changes = []
counter = 0
for row in range(2, 401):
    species_name = LPI_sheet.cell(row=row, column=2).value
#replace "eol_download_2379.csv" with whatever eol file you're using 
    with open("eol_download_2379.csv", "rb") as f:
        reader = csv.reader(f)
        for eol_row in reader:
            test_species_name = eol_row[1].decode('utf-8')
            if test_species_name == species_name:
                counter += 1
                body_size = eol_row[4]
                body_size = body_size.replace(",","")
                body_size = float(body_size)
                measurement_unit = eol_row[7]
#comment this code in and the other set of ifs out if you're working with body mass
                if measurement_unit == "kg":
                    body_size = body_size * 1000
                elif measurement_unit == "log10 grams":
                    body_size = 10**body_size
                elif measurement_unit != "g":
                    break   
#                if measurement_unit == "cm":
#                    body_size = body_size * 10
#                elif measurement_unit == "m":
#                    body_size = body_size * 1000
#                elif measurement_unit == "inch":
#                    body_size = body_size * 25.4
#                elif measurement_unit != "mm":
#                    break  
                body_sizes.append(body_size)
                column = 3
                years = []
                abundances = []
                while (LPI_sheet.cell(row=row, column=column).value != None):
                    year = LPI_sheet.cell(row=row, column=column).value
                    abundance = LPI_sheet.cell(row=row,column=column+1).value
                    years.append(float(year))
                    abundances.append(float(abundance))
                    column = column + 2
                annual_percent_changes = []
                mean_abundances = mean(abundances)
                for num in range(1, len(abundances)-1):
                    difference = abundances[num] - abundances[num-1]
                    percent_change = difference / mean_abundances * 100
                    years_between = years[num] - years[num-1]
                    annual_percent_change = percent_change / years_between
                    annual_percent_changes.append(annual_percent_change)
                average_changes.append(mean(annual_percent_changes))
                break

print stats.linregress(body_sizes, average_changes)                    
               
